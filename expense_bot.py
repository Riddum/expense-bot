#!/usr/bin/env python3
"""
Ledger Bot – with Natural Language, Inline Buttons, and Web App Dashboard.
Optimized for Render deployment with multiprocessing.
"""
import os
import re
import csv
import json
import logging
import asyncio
import psycopg2
import psycopg2.extras
from io import StringIO
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict

import gspread
from oauth2client.service_account import ServiceAccountCredentials
from flask import Flask, request, jsonify
from flask_cors import CORS
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, WebAppInfo
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, filters
)

# ─── CONFIG ──────────────────────────────────────────────────
BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
if not BOT_TOKEN:
    raise SystemExit("Set TELEGRAM_BOT_TOKEN environment variable.")

DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    raise SystemExit("Set DATABASE_URL environment variable (e.g. a free Neon Postgres connection string).")

WEBAPP_URL = os.environ.get("WEBAPP_URL", "https://your-webapp-url.netlify.app")
FLASK_PORT = int(os.environ.get("PORT", 5000))
SYNC_INTERVAL = 60

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ─── CATEGORISATION ────────────────────────────────────────
CATEGORY_RULES = [
    (re.compile(r"rent", re.I), "Housing"),
    (re.compile(r"security", re.I), "Housing"),
    (re.compile(r"\bbb\b|big ?basket", re.I), "Groceries"),
    (re.compile(r"milk", re.I), "Groceries"),
    (re.compile(r"ice ?cream|sweet|snack", re.I), "Snacks & Treats"),
    (re.compile(r"soda|cold ?drink", re.I), "Snacks & Treats"),
    (re.compile(r"home ?trip", re.I), "Family Travel"),
    (re.compile(r"flipkart|myntra|amazon", re.I), "Shopping"),
    (re.compile(r"swiggy|zomato", re.I), "Food Delivery"),
    (re.compile(r"rapido|travel|cab|auto|uber|ola", re.I), "Transport"),
    (re.compile(r"lock", re.I), "Household"),
    (re.compile(r"refreshment", re.I), "Food"),
]

def categorize(note: str) -> str:
    for pattern, cat in CATEGORY_RULES:
        if pattern.search(note):
            return cat
    return "Other"

ALL_CATEGORIES = sorted({cat for _, cat in CATEGORY_RULES} | {"Other"})
CATEGORY_LOOKUP = {c.lower(): c for c in ALL_CATEGORIES}

# ─── MONTH PARSING ───────────────────────────────────────────
MONTHS = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}

def resolve_month_date(month_num: int) -> str:
    """Turn a bare month number into an ISO date, assuming the most recent
    occurrence of that month (this year if it hasn't passed yet, else last
    year). Day is fixed to the 1st since no specific day is given."""
    today = date.today()
    year = today.year if month_num <= today.month else today.year - 1
    return date(year, month_num, 1).isoformat()

# ─── DATABASE ──────────────────────────────────────────────
def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS expenses (
                id SERIAL PRIMARY KEY,
                user_id BIGINT NOT NULL,
                date TEXT NOT NULL,
                note TEXT NOT NULL,
                amount REAL NOT NULL,
                category TEXT NOT NULL,
                user_category TEXT,
                sheet_row INTEGER,
                event TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_settings (
                user_id BIGINT PRIMARY KEY,
                sheet_url TEXT
            )
        """)
        cur.execute("ALTER TABLE expenses ADD COLUMN IF NOT EXISTS event TEXT")
    conn.commit()
    return conn

def get_sheet_url(user_id: int) -> Optional[str]:
    conn = get_db()
    with conn.cursor() as cur:
        cur.execute("SELECT sheet_url FROM user_settings WHERE user_id=%s", (user_id,))
        row = cur.fetchone()
    conn.close()
    return row[0] if row else None

def set_sheet_url(user_id: int, url: str):
    conn = get_db()
    with conn.cursor() as cur:
        cur.execute(
            "INSERT INTO user_settings (user_id, sheet_url) VALUES (%s,%s) "
            "ON CONFLICT (user_id) DO UPDATE SET sheet_url = EXCLUDED.sheet_url",
            (user_id, url)
        )
    conn.commit()
    conn.close()

def add_expense(user_id: int, note: str, amount: float, when: str = None, user_category: str = None, event: str = None) -> str:
    when = when or date.today().isoformat()
    auto_cat = categorize(note)
    final_cat = user_category if user_category else auto_cat
    conn = get_db()
    with conn.cursor() as cur:
        cur.execute(
            "INSERT INTO expenses (user_id, date, note, amount, category, user_category, event) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (user_id, when, note, amount, auto_cat, user_category, event)
        )
        new_id = cur.fetchone()[0]
    conn.commit()
    conn.close()
    # DO NOT SYNC HERE – let periodic sync handle it to avoid blocking
    return final_cat

def get_transactions(user_id: int, start_date: str = None, end_date: str = None) -> List[Dict]:
    conn = get_db()
    query = "SELECT id, date, note, amount, category, user_category, event FROM expenses WHERE user_id=%s"
    params = [user_id]
    if start_date:
        query += " AND date >= %s"
        params.append(start_date)
    if end_date:
        query += " AND date <= %s"
        params.append(end_date)
    query += " ORDER BY date DESC"
    with conn.cursor() as cur:
        cur.execute(query, params)
        rows = cur.fetchall()
    conn.close()
    return [{
        "id": r[0],
        "date": r[1],
        "note": r[2],
        "amount": r[3],
        "category": r[5] if r[5] else r[4],
        "auto_category": r[4],
        "event": r[6]
    } for r in rows]

def delete_expense(user_id: int, tx_id: int):
    conn = get_db()
    with conn.cursor() as cur:
        cur.execute("DELETE FROM expenses WHERE id=%s AND user_id=%s", (tx_id, user_id))
    conn.commit()
    conn.close()

def update_user_category(user_id: int, tx_id: int, new_category: str):
    conn = get_db()
    with conn.cursor() as cur:
        cur.execute("UPDATE expenses SET user_category=%s WHERE id=%s AND user_id=%s", (new_category, tx_id, user_id))
    conn.commit()
    conn.close()

def fmt_inr(n: float) -> str:
    return f"₹{n:,.0f}"

def get_period_data(user_id: int, period: str):
    """Return transactions and total for period: today, week, month."""
    today = date.today()
    if period == "today":
        start = end = today.isoformat()
    elif period == "week":
        start = (today - timedelta(days=today.weekday())).isoformat()
        end = today.isoformat()
    elif period == "month":
        start = today.replace(day=1).isoformat()
        end = today.isoformat()
    else:
        return [], 0
    rows = get_transactions(user_id, start_date=start, end_date=end)
    total = sum(r["amount"] for r in rows)
    return rows, total

# ─── GOOGLE SHEETS (optional) ─────────────────────────────
def get_sheet_client():
    creds_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if creds_json:
        creds_dict = json.loads(creds_json)
        creds = ServiceAccountCredentials.from_json_keyfile_dict(
            creds_dict,
            ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        )
    else:
        # fallback to file (local dev only)
        creds_file = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "service_account.json")
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_name(
                creds_file,
                ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            )
        except FileNotFoundError:
            logger.warning("Google service account credentials not found. Sheet sync disabled.")
            return None
    return gspread.authorize(creds)

def open_sheet(url):
    client = get_sheet_client()
    if client is None:
        raise RuntimeError("No Google Sheets client")
    return client.open_by_url(url).sheet1

def append_to_sheet(url, row):
    sheet = open_sheet(url)
    sheet.append_row(row)

def sync_sheet_to_db(user_id: int):
    url = get_sheet_url(user_id)
    if not url:
        return
    try:
        sheet = open_sheet(url)
        rows = sheet.get_all_values()
        if not rows:
            return
        conn = get_db()
        cur = conn.cursor()
        for i, row in enumerate(rows[1:], start=2):
            if len(row) < 4:
                continue
            date_str, note, amt_str, cat = row[:4]
            try:
                amount = float(amt_str)
            except ValueError:
                continue
            # check if already exists (by date, note, amount)
            cur.execute(
                "SELECT id, sheet_row FROM expenses WHERE user_id=%s AND date=%s AND note=%s AND amount=%s",
                (user_id, date_str, note, amount)
            )
            existing = cur.fetchone()
            if existing:
                if not existing[1]:
                    cur.execute("UPDATE expenses SET sheet_row=%s WHERE id=%s", (i, existing[0]))
                    conn.commit()
                continue
            auto_cat = categorize(note)
            cur.execute(
                "INSERT INTO expenses (user_id, date, note, amount, category, user_category, sheet_row) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (user_id, date_str, note, amount, auto_cat, cat if cat else None, i)
            )
            conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Sync failed for user {user_id}: {e}")

# ─── FLASK API FOR WEB APP ─────────────────────────────────
flask_app = Flask(__name__)
CORS(flask_app)  # enable CORS for all routes

@flask_app.route("/health", methods=["GET"])
def health_check():
    return jsonify({"status": "ok"}), 200

@flask_app.route("/data", methods=["GET"])
def get_data():
    user_id = request.args.get("user_id")
    period = request.args.get("period", "all")
    if not user_id:
        return jsonify({"error": "Missing user_id"}), 400
    try:
        uid = int(user_id)
    except ValueError:
        return jsonify({"error": "Invalid user_id"}), 400

    if period == "today":
        start_date = date.today().isoformat()
    elif period == "week":
        start_date = (date.today() - timedelta(days=date.today().weekday())).isoformat()
    elif period == "month":
        start_date = date.today().replace(day=1).isoformat()
    else:  # "all" or unrecognized -> no lower bound
        start_date = None

    rows = get_transactions(uid, start_date=start_date)
    total = sum(r["amount"] for r in rows)

    cat_totals = {}
    for r in rows:
        cat_totals[r["category"]] = cat_totals.get(r["category"], 0) + r["amount"]

    return jsonify({
        "transactions": rows,
        "total": total,
        "categories": cat_totals,
        "chartData": {
            "labels": list(cat_totals.keys()),
            "values": list(cat_totals.values())
        }
    })

# ─── TELEGRAM HANDLERS ─────────────────────────────────────
def make_keyboard():
    buttons = [
        [InlineKeyboardButton("📊 Today", callback_data="report_today"),
         InlineKeyboardButton("📅 Week", callback_data="report_week"),
         InlineKeyboardButton("📆 Month", callback_data="report_month"),
         InlineKeyboardButton("📂 Categories", callback_data="report_categories")],
        [InlineKeyboardButton("🌐 Open Dashboard", web_app=WebAppInfo(url=WEBAPP_URL))]
    ]
    return InlineKeyboardMarkup(buttons)

async def send_report(update: Update, context: ContextTypes.DEFAULT_TYPE, period: str):
    user_id = update.effective_user.id
    rows, total = get_period_data(user_id, period)
    if not rows:
        msg = f"No expenses for {period}."
    else:
        lines = [
            f"• {r['note']}" + (f" [{r['event']}]" if r['event'] else "") + f" — {fmt_inr(r['amount'])} ({r['category']})"
            for r in rows[:10]
        ]
        more = f"\n... and {len(rows)-10} more" if len(rows) > 10 else ""
        msg = f"📊 {period.capitalize()} total: {fmt_inr(total)}\n\n" + "\n".join(lines) + more
    await update.callback_query.answer()
    await update.callback_query.edit_message_text(msg, reply_markup=make_keyboard())

async def report_categories(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    month_start = date.today().replace(day=1).isoformat()
    rows = get_transactions(user_id, start_date=month_start)
    if not rows:
        msg = "No expenses this month."
    else:
        cat_totals = {}
        for r in rows:
            cat_totals[r["category"]] = cat_totals.get(r["category"], 0) + r["amount"]
        sorted_cats = sorted(cat_totals.items(), key=lambda x: -x[1])
        lines = [f"• {cat}: {fmt_inr(amt)}" for cat, amt in sorted_cats]
        total = sum(cat_totals.values())
        msg = f"📂 Category breakdown (this month): {fmt_inr(total)}\n\n" + "\n".join(lines)
    await update.callback_query.answer()
    await update.callback_query.edit_message_text(msg, reply_markup=make_keyboard())

async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    if data.startswith("report_"):
        period = data.split("_")[1]
        if period == "categories":
            await report_categories(update, context)
        else:
            await send_report(update, context, period)
    else:
        await query.answer("Unknown action")

# ─── NATURAL LANGUAGE PARSING ─────────────────────────────
# Supported shapes (amount can be first, last, or in the middle):
#   "150 bb"                -> amount=150, note=bb
#   "bb 150"                -> amount=150, note=bb
#   "rent 17000"            -> amount=17000, note=rent
#   "ice cream 200 snacks"  -> note="ice cream", amount=200, event="snacks"
#   "150 bb may"            -> note=bb, amount=150, dated within May
#   "150 bb interview may"  -> note=bb, amount=150, event="interview", dated within May
def parse_message(text: str):
    """Returns (amount, note, event, month_num) or None."""
    tokens = text.strip().split()
    if not tokens:
        return None

    month_num = None
    if tokens[-1].lower() in MONTHS:
        month_num = MONTHS[tokens[-1].lower()]
        tokens = tokens[:-1]
    if not tokens:
        return None

    amount_idx = None
    for i, tok in enumerate(tokens):
        try:
            float(tok)
            amount_idx = i
            break
        except ValueError:
            continue
    if amount_idx is None:
        return None

    amount = float(tokens[amount_idx])
    before = tokens[:amount_idx]
    after = tokens[amount_idx + 1:]

    if before:
        note = " ".join(before)
        event = " ".join(after) if after else None
    elif after:
        note = after[0]
        event = " ".join(after[1:]) if len(after) > 1 else None
    else:
        return None

    return amount, note, event, month_num

async def handle_plain_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    parsed = parse_message(text)
    if not parsed:
        await update.message.reply_text("Sorry, I didn't understand. Try: 'bb 150' or '150 bb'.")
        return
    amount, note, event, month_num = parsed
    user_id = update.effective_user.id
    when = resolve_month_date(month_num) if month_num else None
    final_cat = add_expense(user_id, note, amount, when=when, event=event)
    reply = f"✅ Logged {fmt_inr(amount)} · {note} → {final_cat}"
    if event:
        reply += f" 🏷️ {event}"
    if when:
        reply += f" ({when})"
    await update.message.reply_text(reply, reply_markup=make_keyboard())

async def cmd_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = " ".join(context.args)
    if not text:
        await update.message.reply_text("Usage: /add 150 bb")
        return
    parsed = parse_message(text)
    if not parsed:
        await update.message.reply_text("Invalid format. Try: /add 150 bb")
        return
    amount, note, event, month_num = parsed
    user_id = update.effective_user.id
    when = resolve_month_date(month_num) if month_num else None
    final_cat = add_expense(user_id, note, amount, when=when, event=event)
    reply = f"✅ Logged {fmt_inr(amount)} · {note} → {final_cat}"
    if event:
        reply += f" 🏷️ {event}"
    if when:
        reply += f" ({when})"
    await update.message.reply_text(reply, reply_markup=make_keyboard())

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📒 Ledger Bot\n\n"
        "Just type something like:\n"
        "  • bb 150\n"
        "  • 150 rapido\n"
        "  • rent 17000\n\n"
        "Use the buttons below for reports and dashboard.",
        reply_markup=make_keyboard()
    )

async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Commands:\n"
        "/start – welcome\n"
        "/help – this help\n"
        "/delete <id> – remove a transaction\n"
        "/setcat <id> <category> – change category\n"
        "/listcats – show all categories\n"
        "/setsheet <url> – link Google Sheet\n"
        "/export – CSV export\n\n"
        "Add expenses by just typing, e.g.:\n"
        "  • bb 150\n"
        "  • 150 rapido\n"
        "  • 150 bb may – logs it dated within May\n"
        "  • 150 bb interview may – tags it with the 'interview' "
        "event, still counted in May's total\n\n"
        "📎 Send an .xlsx file directly to bulk-import expenses "
        "(first column = note, other columns = month names or event names)."
    )

async def cmd_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Usage: /delete <id>")
        return
    try:
        tx_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("Invalid ID.")
        return
    delete_expense(update.effective_user.id, tx_id)
    await update.message.reply_text(f"Deleted #{tx_id}.")

async def cmd_setcat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text("Usage: /setcat <id> <category>")
        return
    try:
        tx_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("Invalid ID.")
        return
    new_cat = " ".join(context.args[1:])
    if new_cat not in ALL_CATEGORIES:
        await update.message.reply_text(f"Category must be one of: " + ", ".join(ALL_CATEGORIES))
        return
    update_user_category(update.effective_user.id, tx_id, new_cat)
    await update.message.reply_text(f"Category for #{tx_id} updated to '{new_cat}'.")

async def cmd_listcats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Available categories:\n" + ", ".join(ALL_CATEGORIES))

async def cmd_setsheet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Usage: /setsheet <google_sheet_url>")
        return
    url = context.args[0].strip()
    if not url.startswith("https://docs.google.com/spreadsheets"):
        await update.message.reply_text("Please provide a valid Google Sheets URL.")
        return
    set_sheet_url(update.effective_user.id, url)
    await update.message.reply_text("Google Sheet linked! Syncing will start shortly.")
    try:
        sync_sheet_to_db(update.effective_user.id)
        await update.message.reply_text("Initial sync completed.")
    except Exception as e:
        await update.message.reply_text(f"Sync error: {e}")

async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    rows = get_transactions(user_id)
    if not rows:
        await update.message.reply_text("Nothing to export.")
        return
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(["date", "note", "amount", "category", "event"])
    for r in rows:
        writer.writerow([r["date"], r["note"], r["amount"], r["category"], r["event"] or ""])
    buf.seek(0)
    await update.message.reply_document(
        document=buf.getvalue().encode(),
        filename="expenses.csv",
        caption="All transactions."
    )

# ─── XLSX IMPORT ────────────────────────────────────────────
# Bulk-import a spreadsheet shaped like: first column = expense note,
# other columns = month names ("may", "june", ...) each holding amounts,
# plus optional "event" columns (e.g. "interview day") whose amounts get
# tagged with that event and dated using whichever month column has data
# in the same row (or the nearest preceding month column as a fallback).
def _nearest_preceding_month(col_info, idx):
    for i in range(idx - 1, -1, -1):
        if col_info[i][0] == "month":
            return col_info[i][1]
    return None

def import_expenses_from_xlsx(user_id: int, file_bytes: bytes):
    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0, 0

    header = rows[0]
    col_info = []  # (kind, meta) per column: kind in {"note","month","event","skip"}
    for idx, h in enumerate(header):
        if idx == 0:
            col_info.append(("note", None))
            continue
        if h is None or not str(h).strip():
            col_info.append(("skip", None))
            continue
        h_str = str(h).strip().lower()
        if h_str in MONTHS:
            col_info.append(("month", MONTHS[h_str]))
        else:
            event_tag = re.sub(r"\s*(day|date)\s*$", "", h_str).strip() or h_str
            col_info.append(("event", event_tag))

    imported, skipped = 0, 0
    for row in rows[1:]:
        note = row[0] if row else None
        if not note or not str(note).strip():
            skipped += 1
            continue
        note = str(note).strip()

        row_month = None
        for idx, val in enumerate(row):
            if idx < len(col_info) and col_info[idx][0] == "month" and val not in (None, ""):
                row_month = col_info[idx][1]
                break

        for idx, val in enumerate(row):
            if idx == 0 or idx >= len(col_info) or val in (None, ""):
                continue
            kind, meta = col_info[idx]
            if kind == "month":
                try:
                    amount = float(val)
                except (TypeError, ValueError):
                    continue
                add_expense(user_id, note, amount, when=resolve_month_date(meta))
                imported += 1
            elif kind == "event":
                try:
                    amount = float(val)
                except (TypeError, ValueError):
                    continue
                month_num = row_month or _nearest_preceding_month(col_info, idx)
                when = resolve_month_date(month_num) if month_num else date.today().isoformat()
                add_expense(user_id, note, amount, when=when, event=meta)
                imported += 1

    return imported, skipped

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc.file_name.lower().endswith((".xlsx", ".xls")):
        await update.message.reply_text("Send an .xlsx file to import expenses.")
        return
    await update.message.reply_text("📥 Importing... this may take a moment.")
    user_id = update.effective_user.id
    try:
        file = await doc.get_file()
        file_bytes = bytes(await file.download_as_bytearray())
        imported, skipped = import_expenses_from_xlsx(user_id, file_bytes)
    except Exception as e:
        logger.error(f"Import failed for {user_id}: {e}")
        await update.message.reply_text(f"❌ Import failed: {e}")
        return
    msg = f"✅ Imported {imported} transaction(s)."
    if skipped:
        msg += f" Skipped {skipped} row(s) with no note (e.g. totals rows)."
    msg += "\n\nReview with the buttons below, or use /export to double check."
    await update.message.reply_text(msg, reply_markup=make_keyboard())

async def periodic_sync(application):
    """Background task to sync with Google Sheets periodically"""
    while True:
        try:
            await asyncio.sleep(SYNC_INTERVAL)
            conn = get_db()
            with conn.cursor() as cur:
                cur.execute("SELECT user_id FROM user_settings WHERE sheet_url IS NOT NULL")
                users = cur.fetchall()
            conn.close()
            for (uid,) in users:
                try:
                    sync_sheet_to_db(uid)
                except Exception as e:
                    logger.error(f"Sync error for {uid}: {e}")
        except Exception as e:
            logger.error(f"Periodic sync error: {e}")

# ─── TELEGRAM BOT ENTRY ─────────────────────────────────────
async def _post_init(app):
    """PTB calls this once, after initialize() but before polling starts.
    This is the correct place to schedule background asyncio tasks --
    run_polling() owns the event loop, so we can't create tasks before it."""
    asyncio.create_task(periodic_sync(app))

def main_telegram():
    """Run the Telegram bot.
    IMPORTANT: run_polling() is a BLOCKING call that creates and manages its
    own event loop internally. It must be called directly, never awaited
    inside asyncio.run()/another coroutine -- doing so causes PTB to crash
    with 'RuntimeError: Cannot close a running event loop' once it tries to
    tear down a loop that asyncio.run() is still managing."""
    app = ApplicationBuilder().token(BOT_TOKEN).post_init(_post_init).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("add", cmd_add))
    app.add_handler(CommandHandler("delete", cmd_delete))
    app.add_handler(CommandHandler("setcat", cmd_setcat))
    app.add_handler(CommandHandler("listcats", cmd_listcats))
    app.add_handler(CommandHandler("setsheet", cmd_setsheet))
    app.add_handler(CommandHandler("export", cmd_export))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_plain_message))
    app.add_handler(CallbackQueryHandler(callback_handler))

    logger.info("Telegram bot started and listening for updates...")
    app.run_polling()

def run_telegram_bot():
    """Entry point for Telegram bot in a separate process"""
    try:
        main_telegram()
    except KeyboardInterrupt:
        logger.info("Bot interrupted")
